import re
from typing import Dict, List, Tuple, Optional, Union
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from table_structure import *


def is_number(num: str) -> bool:
    pattern = re.compile(r'^\s*[-+]?(\d+\.?\d*|\d*\.?\d+)%?\s*$')
    result = pattern.match(num)
    if result:
        return True
    else:
        return False

def to_number(num: Union[str, float, int]) -> Optional[float]:
    if isinstance(num, float) or isinstance(num, int) or num is None:
        return num
    if is_number(num):
        num = num.strip()
        if num.endswith('%'):
            return float(num[:-1]) / 100
        else:
            return float(num)
    else:
        return None

def mine_content_process(ws: Worksheet, cover: bool) -> Tuple[Dict[str, MineContentInfo], int]:
    cover_num = 0
    instance_map = {}
    head_map = {
        'id': 0,
        'clay': 1,
        'quartz': 2,
        'sand_quartz': 3,
        'sand_feldspar': 4,
        'sand_other': 5,
        'debris': 7,
        'hollow_close': 8,
        'hollow_open': 9,
        'hollow_through': 10
    }
    for i, row in enumerate(ws.iter_rows(3, ws.max_row, 1, 11, values_only=True)):
        add_flag = True
        instance = MineContentInfo()
        instance.id = str(row[head_map['id']]).strip()
        ins = MineContentInfo.query.filter_by(id=instance.id).first()
        if ins:
            if cover:
                instance = ins
                cover_num += 1
                add_flag = False
            else:
                continue
        instance.clay = to_number(row[head_map['clay']])
        instance.quartz = to_number(row[head_map['quartz']])
        instance.sand_quartz = to_number(row[head_map['sand_quartz']])
        instance.sand_feldspar = to_number(row[head_map['sand_feldspar']])
        instance.sand_other = to_number(row[head_map['sand_other']])
        instance.debris = to_number(row[head_map['debris']])
        instance.hollow_close = to_number(row[head_map['hollow_close']])
        instance.hollow_open = to_number(row[head_map['hollow_open']])
        instance.hollow_through = to_number(row[head_map['hollow_through']])
        if add_flag:
            instance_map[instance.id] = instance
    return instance_map, cover_num

def mine_survey_process(ws: Worksheet, cover: bool) -> Tuple[Dict[str, MineSurveyInfo], int]:
    cover_num = 0
    instance_map = {}
    head_map = {
        'id': 0,
        'debris_0um': 1,
        'debris_67um': 2,
        'debris_167um': 3,
        'debris_501um': 4,
        'debris_1002um': 5,
        'hollow_0um': 6,
        'hollow_67um': 7,
        'hollow_501um': 8,
        'hollow_1002um': 9,
        'hollow_2004um': 10
    }
    for i, row in enumerate(ws.iter_rows(3, ws.max_row, 1, 11, values_only=True)):
        add_flag = True
        instance = MineSurveyInfo()
        instance.id = str(row[head_map['id']]).strip()
        ins = MineSurveyInfo.query.filter_by(id=instance.id).first()
        if ins:
            if cover:
                instance = ins
                cover_num += 1
                add_flag = False
            else:
                continue
        instance.debris_0um = to_number(row[head_map['debris_0um']])
        instance.debris_67um = to_number(row[head_map['debris_67um']])
        instance.debris_167um = to_number(row[head_map['debris_167um']])
        instance.debris_501um = to_number(row[head_map['debris_501um']])
        instance.debris_1002um = to_number(row[head_map['debris_1002um']])
        instance.hollow_0um = to_number(row[head_map['hollow_0um']])
        instance.hollow_67um = to_number(row[head_map['hollow_67um']])
        instance.hollow_501um = to_number(row[head_map['hollow_501um']])
        instance.hollow_1002um = to_number(row[head_map['hollow_1002um']])
        instance.hollow_2004um = to_number(row[head_map['hollow_2004um']])
        if add_flag:
            instance_map[instance.id] = instance
    return instance_map, cover_num

def mine_xrd_process(ws: Worksheet, cover: bool) -> Tuple[Dict[str, MineXRDInfo], int]:
    cover_num = 0
    instance_map = {}
    head_map = {
        'id': 0,
        'type': 1,
        'quartz': 2,
        'albite': 3,
        'potashFeldspar': 4,
        'mica': 5,
        'amphibole': 6,
        'hematite': 7,
        'magnetite': 8,
        'dolomite': 9,
        'analcite': 10,
        'tridymite': 11,
        'cristobalite': 12,
        'mullite': 13,
        'other': 14
    }
    for i, row in enumerate(ws.iter_rows(3, ws.max_row, 1, 15, values_only=True)):
        add_flag = True
        instance = MineXRDInfo()
        instance.id = str(row[head_map['id']]).strip()
        ins = MineXRDInfo.query.filter_by(id=instance.id).first()
        if ins:
            if cover:
                instance = ins
                cover_num += 1
                add_flag = False
            else:
                continue
        t = row[head_map['type']].strip()
        instance.type = t if t else None
        instance.quartz = to_number(row[head_map['quartz']])
        instance.albite = to_number(row[head_map['albite']])
        instance.potashFeldspar = to_number(row[head_map['potashFeldspar']])
        instance.mica = to_number(row[head_map['mica']])
        instance.amphibole = to_number(row[head_map['amphibole']])
        instance.hematite = to_number(row[head_map['hematite']])
        instance.magnetite = to_number(row[head_map['magnetite']])
        instance.dolomite = to_number(row[head_map['dolomite']])
        instance.analcite = to_number(row[head_map['analcite']])
        instance.tridymite = to_number(row[head_map['tridymite']])
        instance.cristobalite = to_number(row[head_map['cristobalite']])
        instance.mullite = to_number(row[head_map['mullite']])
        instance.other = to_number(row[head_map['other']])
        if add_flag:
            instance_map[instance.id] = instance
    return instance_map, cover_num

def mine_chemistry_process(ws: Worksheet, cover: bool) -> Tuple[Dict[str, MineChemistryInfo], int]:
    cover_num = 0
    instance_map = {}
    head_map = {
        'id': 0,
        'Na2O': 1,
        'MgO': 2,
        'Al2O3': 3,
        'SiO2': 4,
        'P2O5': 5,
        'SO2': 6,
        'K2O': 7,
        'CaO': 8,
        'TiO2': 9,
        'MnO': 10,
        'FeO': 11,
        'CuO': 12,
        'ZnO': 13,
        'As2O3': 14,
        'SnO2': 15,
        'PbO': 16,
        'other': 17
    }
    for i, row in enumerate(ws.iter_rows(3, ws.max_row, 1, 18, values_only=True)):
        add_flag = True
        instance = MineChemistryInfo()
        instance.id = str(row[head_map['id']]).strip()
        ins = MineChemistryInfo.query.filter_by(id=instance.id).first()
        if ins:
            if cover:
                instance = ins
                cover_num += 1
                add_flag = False
            else:
                continue
        instance.Na2O = to_number(row[head_map['Na2O']])
        instance.MgO = to_number(row[head_map['MgO']])
        instance.Al2O3 = to_number(row[head_map['Al2O3']])
        instance.SiO2 = to_number(row[head_map['SiO2']])
        instance.P2O5 = to_number(row[head_map['P2O5']])
        instance.SO2 = to_number(row[head_map['SO2']])
        instance.K2O = to_number(row[head_map['K2O']])
        instance.CaO = to_number(row[head_map['CaO']])
        instance.TiO2 = to_number(row[head_map['TiO2']])
        instance.MnO = to_number(row[head_map['MnO']])
        instance.FeO = to_number(row[head_map['FeO']])
        instance.CuO = to_number(row[head_map['CuO']])
        instance.ZnO = to_number(row[head_map['ZnO']])
        instance.As2O3 = to_number(row[head_map['As2O3']])
        instance.SnO2 = to_number(row[head_map['SnO2']])
        instance.PbO = to_number(row[head_map['PbO']])
        instance.other = to_number(row[head_map['other']])
        if add_flag:
            instance_map[instance.id] = instance
    return instance_map, cover_num

def mine_chemistry_single_process(ws: Worksheet, cover: bool) -> Tuple[Dict[str, MineChemistryInfoSingle], int]:
    cover_num = 0
    instance_map = {}
    head_map = {
        'id': 0,
        'C': 1,
        'Na': 2,
        'Mg': 3,
        'Al': 4,
        'Si': 5,
        'P': 6,
        'S': 7,
        'Cl': 8,
        'K': 9,
        'Ca': 10,
        'Ti': 11,
        'V': 12,
        'Mn': 13,
        'Fe': 14,
        'Co': 15,
        'Ni': 16,
        'Cu': 17,
        'Zn': 18,
        'As': 19,
        'Ag': 20,
        'Sn': 21,
        'Sb': 22,
        'Au': 23,
        'Hg': 24,
        'Pb': 25,
        'other': 26
    }
    for i, row in enumerate(ws.iter_rows(3, ws.max_row, 1, 27, values_only=True)):
        add_flag = True
        instance = MineChemistryInfoSingle()
        instance.id = str(row[head_map['id']]).strip()
        ins = MineChemistryInfoSingle.query.filter_by(id=instance.id).first()
        if ins:
            if cover:
                instance = ins
                cover_num += 1
                add_flag = False
            else:
                continue
        instance.C = to_number(row[head_map['C']])
        instance.Na = to_number(row[head_map['Na']])
        instance.Mg = to_number(row[head_map['Mg']])
        instance.Al = to_number(row[head_map['Al']])
        instance.Si = to_number(row[head_map['Si']])
        instance.P = to_number(row[head_map['P']])
        instance.S = to_number(row[head_map['S']])
        instance.Cl = to_number(row[head_map['Cl']])
        instance.K = to_number(row[head_map['K']])
        instance.Ca = to_number(row[head_map['Ca']])
        instance.Ti = to_number(row[head_map['Ti']])
        instance.V = to_number(row[head_map['V']])
        instance.Mn = to_number(row[head_map['Mn']])
        instance.Fe = to_number(row[head_map['Fe']])
        instance.Co = to_number(row[head_map['Co']])
        instance.Ni = to_number(row[head_map['Ni']])
        instance.Cu = to_number(row[head_map['Cu']])
        instance.Zn = to_number(row[head_map['Zn']])
        instance.As = to_number(row[head_map['As']])
        instance.Ag = to_number(row[head_map['Ag']])
        instance.Sn = to_number(row[head_map['Sn']])
        instance.Sb = to_number(row[head_map['Sb']])
        instance.Au = to_number(row[head_map['Au']])
        instance.Hg = to_number(row[head_map['Hg']])
        instance.Pb = to_number(row[head_map['Pb']])
        instance.other = to_number(row[head_map['other']])
        if add_flag:
            instance_map[instance.id] = instance
    return instance_map, cover_num

def mine_physic_process(ws: Worksheet, cover: bool) -> Tuple[Dict[str, MinePhysicsInfo], int]:
    cover_num = 0
    instance_map = {}
    head_map = {
        'id': 0,
        'type': 1,
        'trueDensity': 2,
        'apparentPorosity': 3,
        'waterAbsorption': 4,
        'bending': 5,
        'resistance': 6,
        'slag': 7,
        'alkali': 8,
        'refractoriness': 9,
        'heat': 10
    }
    for i, row in enumerate(ws.iter_rows(2, ws.max_row, 1, 11, values_only=True)):
        add_flag = True
        instance = MinePhysicsInfo()
        instance.id = str(row[head_map['id']]).strip()
        ins = MinePhysicsInfo.query.filter_by(id=instance.id).first()
        if ins:
            if cover:
                instance = ins
                cover_num += 1
                add_flag = False
            else:
                continue
        t = row[head_map['type']].strip()
        instance.type = t if t else None
        instance.trueDensity = to_number(row[head_map['trueDensity']])
        instance.apparentPorosity = to_number(row[head_map['apparentPorosity']])
        instance.waterAbsorption = to_number(row[head_map['waterAbsorption']])
        instance.bending = to_number(row[head_map['bending']])
        instance.resistance = to_number(row[head_map['resistance']])
        instance.slag = to_number(row[head_map['slag']])
        instance.alkali = to_number(row[head_map['alkali']])
        instance.refractoriness = to_number(row[head_map['refractoriness']])
        instance.heat = to_number(row[head_map['heat']])
        if add_flag:
            instance_map[instance.id] = instance
    return instance_map, cover_num

def mine_thermal_process(ws: Worksheet, cover: bool) -> Tuple[Dict[str, MineThermalInfo], int]:
    cover_num = 0
    instance_map = {}
    head_map = {
        'id': 0,
        'melting': 1,
        'fireResis': 2,
        'termTemper': 3
        # 'data': 4,
        # 'surveImage': 5
    }
    for i, row in enumerate(ws.iter_rows(2, ws.max_row, 1, 4, values_only=True)):
        add_flag = True
        instance = MineThermalInfo()
        instance.id = str(row[head_map['id']]).strip()
        ins = MineThermalInfo.query.filter_by(id=instance.id).first()
        if ins:
            if cover:
                instance = ins
                cover_num += 1
                add_flag = False
            else:
                continue
        instance.melting = to_number(row[head_map['melting']])
        instance.fireResis = to_number(row[head_map['fireResis']])
        instance.termTemper = to_number(row[head_map['termTemper']])
        # instance.data = row[head_map['data']].strip()
        # instance.surveImage = row[head_map['surveImage']].strip()
        if add_flag:
            instance_map[instance.id] = instance
    return instance_map, cover_num

def get_instances(wb: Workbook, cover: bool) -> Tuple[List, List[SampleInfo], int]:
    cover_num = 0
    sample_set = set()
    sample_type_dict = {}
    instance_list = []
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        sheetname = sheetname.strip()
        if sheetname == '矿物含量':
            il, con = mine_content_process(ws, cover)
            instance_list.extend(il.values())
            sample_set.update(il.keys())
            cover_num += con
        elif sheetname == '矿物测量':
            il, con = mine_survey_process(ws, cover)
            instance_list.extend(il.values())
            sample_set.update(il.keys())
            cover_num += con
        elif sheetname == 'XRD':
            il, con = mine_xrd_process(ws, cover)
            for k, v in il.items():
                sample_type_dict[k] = v.type
            instance_list.extend(il.values())
            sample_set.update(il.keys())
            cover_num += con
        elif sheetname == '化学成分(氧化物)':
            il, con = mine_chemistry_process(ws, cover)
            instance_list.extend(il.values())
            sample_set.update(il.keys())
            cover_num += con
        elif sheetname == '化学成分(单质)':
            il, con = mine_chemistry_single_process(ws, cover)
            instance_list.extend(il.values())
            sample_set.update(il.keys())
            cover_num += con
        elif sheetname == '物理性能':
            il, con = mine_physic_process(ws, cover)
            for k, v in il.items():
                sample_type_dict[k] = v.type
            instance_list.extend(il.values())
            sample_set.update(il.keys())
            cover_num += con
        elif sheetname == '热分析':
            il, con = mine_thermal_process(ws, cover)
            instance_list.extend(il.values())
            sample_set.update(il.keys())
            cover_num += con
        else:
            print(f'import_excel ----> 意料之外的sheet：{sheetname}')
    sample_list = [SampleInfo(iid, type=sample_type_dict.get(iid, None)) for iid in sample_set]
    return instance_list, sample_list, cover_num


if __name__ == '__main__':
    print(get_instances(load_workbook('../file/public/excels/导入模板.xlsx', read_only=True), cover=True))
